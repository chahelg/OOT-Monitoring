"""
interface_matrix.py — normalizes Oxygen_Interface_Matrix_w1.xlsx into one
clean InterfaceRecord catalog, and scores alert rows against it to
resolve which interface (and stream) each alert belongs to.

Pure, Flask-free, independently runnable:
    py interface_matrix.py --self-check "output\\active-alert category 19 Aug.xlsx"

Only reads the "Interface Matrix R1+R2" sheet (203 rows, 47 columns) —
the second sheet, "Interface Matrix R2" (31 rows), is intentionally not
read. It was verified to be a redundant enrichment layer over a subset
of R1+R2's own rows (every one of its 31 rows matches an existing
(Application, Interface Name) pair in R1+R2 — confirmed by exact join,
31/31 matched, zero new interfaces there), not a separate source of
interfaces — by request, this module keeps R1+R2 as the single source
rather than merging the two.
"""

from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

APP_DIR = Path(__file__).parent
DEFAULT_MATRIX_PATH = APP_DIR / "reference" / "Oxygen_Interface_Matrix_w1.xlsx"
PRIMARY_SHEET = "Interface Matrix R1+R2"

# Confirmed by inspecting real column values, not assumed: these three
# strings are not application names, they mean "no Datadog app to join
# on" — appear in both sheets, verified via Counter() on the real file.
NOT_DATADOG_VALUES = {
    "No Datadog - only BTP API",
    "No Datadog - only CTM Job",
    "No Datadog - direct RFC Connection",
}

# Real case/whitespace variance confirmed in the file ('low' vs 'Low',
# 'High ' vs 'High') — normalize rather than let it silently split a
# criticality bucket into two.
_CRITICALITY_MAP = {"critical": "Critical", "high": "High", "medium": "Medium", "low": "Low"}


# --------------------------------------------------------------------
# Cell-value cleaning — every raw cell passes through one of these
# before it becomes a field on InterfaceRecord. None means "unknown",
# never a false/zero default — sparse enrichment (e.g. FI impact is
# only filled on 84 of 203 rows) must read as unknown, not "No".
# --------------------------------------------------------------------

def _clean_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _clean_criticality(v) -> str | None:
    s = _clean_str(v)
    if s is None:
        return None
    return _CRITICALITY_MAP.get(s.lower(), s)


def _clean_yesno(v) -> bool | None:
    s = _clean_str(v)
    if s is None:
        return None
    low = s.lower()
    if low in ("yes", "y"):
        return True
    if low in ("no", "n"):
        return False
    return None  # some other free-text value — unknown, not False


def _split_streams(v) -> list[str]:
    s = _clean_str(v)
    if not s:
        return []
    # Sorted, not source order: "SCM-FTM" must attribute to FTM
    # (alphabetically first), not SCM just because it was written
    # first — the brief is explicit that attribution is alphabetical,
    # and primary_stream()/resolve_alert_interface both just take
    # streams[0], so the sort has to happen here to be honored anywhere
    # that reads it.
    return sorted(p.strip() for p in re.split(r"[-+]", s) if p.strip())


def _split_api_identifiers(v) -> list[str]:
    s = _clean_str(v)
    if not s:
        return []
    return [line.strip().strip('"') for line in s.split("\n") if line.strip().strip('"')]


@dataclass
class InterfaceRecord:
    application: str | None = None
    location: str | None = None
    business_category: str | None = None
    streams: list[str] = field(default_factory=list)
    interface_name: str | None = None
    # Not in the brief's field table, but empirically necessary: when
    # two candidate interfaces for one application share the same API
    # identifier (a real, confirmed case — ASTRO-WMS-BENE's "Outbound
    # Delivery" and "Shipment Confirmation(PGI)" both list
    # API_OUTBOUND_DELIVERY_SRV), this free-text description is what
    # actually disambiguates them (e.g. "...Post Good Issue..." only
    # appears on the correct one). Dropping it during the rewrite
    # silently misattributed the single largest real alert category —
    # caught by re-checking against the ground truth established
    # earlier this session, not assumed correct from the field list.
    aif_description: str | None = None
    middleware: str | None = None
    direction: str | None = None
    datadog_app: str | None = None
    not_datadog_monitored: bool = False
    # Always None: only ever populated from "Interface Matrix R2", which
    # this module intentionally doesn't read (see module docstring).
    # Kept on the schema rather than removed in case that changes later.
    datadog_business_object: str | None = None
    api_identifiers: list[str] = field(default_factory=list)
    in_daily_report: bool | None = None
    interface_criticality: str | None = None
    application_criticality: str | None = None
    business_criticality: str | None = None
    time_critical: bool | None = None
    fi_posting_impact: bool | None = None
    stock_movement: bool | None = None  # always None — same reason as datadog_business_object above
    ams_team: str | None = None
    app_owner_it: str | None = None
    key_contacts: str | None = None
    operational_contacts: str | None = None
    core_bpo: str | None = None
    local_bpo: str | None = None
    source_sheets: list[str] = field(default_factory=list)

    def primary_stream(self) -> str | None:
        return self.streams[0] if self.streams else None

    def also_streams(self) -> list[str]:
        return self.streams[1:]


def _sheet_rows(ws) -> tuple[list[str], list[tuple]]:
    rows_iter = ws.iter_rows(values_only=True)
    header = [(_clean_str(h) or "") for h in next(rows_iter)]
    data_rows = [r for r in rows_iter if any(v is not None for v in r)]
    return header, data_rows


def _row_get(header: list[str], row: tuple, col_name: str):
    try:
        idx = header.index(col_name)
    except ValueError:
        raise ValueError(f"Expected column {col_name!r} not found. Found: {header}")
    return row[idx] if idx < len(row) else None


def _record_from_primary_row(header: list[str], row: tuple) -> InterfaceRecord:
    def get(col):
        return _row_get(header, row, col)

    datadog_app_raw = _clean_str(get("Datadog ApplicationName"))
    not_monitored = datadog_app_raw in NOT_DATADOG_VALUES
    return InterfaceRecord(
        application=_clean_str(get("Application")),
        location=_clean_str(get("Location")),
        business_category=_clean_str(get("Category")),
        streams=_split_streams(get("Stream")),
        interface_name=_clean_str(get("Interface Name")),
        aif_description=_clean_str(get("AIF Interface Description")),
        middleware=_clean_str(get("Middleware (Primary)")),
        direction=_clean_str(get("Message Direction")),
        datadog_app=None if not_monitored else datadog_app_raw,
        not_datadog_monitored=not_monitored,
        datadog_business_object=None,  # only ever comes from R2 enrichment
        api_identifiers=_split_api_identifiers(get("API Technical details( SOAP / ODATA)")),
        in_daily_report=_clean_yesno(get("DataDog Daily Report")),
        interface_criticality=_clean_criticality(get("Interface Criticality")),
        application_criticality=_clean_criticality(get("Application Criticality")),
        business_criticality=_clean_criticality(get("Business function Criticality(Per Application)")),
        time_critical=_clean_yesno(get("Time critical (Yes/NO) by Strem aleads")),
        fi_posting_impact=_clean_yesno(get("Impact on FI Posting")),
        stock_movement=None,  # only ever comes from R2 enrichment
        ams_team=_clean_str(get("AMS Team")),
        app_owner_it=_clean_str(get("Application Owner IT")),
        key_contacts=_clean_str(get("Key Contacts")),
        operational_contacts=_clean_str(get("Operational Key Contacts")),
        core_bpo=_clean_str(get("Core BPO")),
        local_bpo=_clean_str(get("Local BPO")),
        source_sheets=[PRIMARY_SHEET],
    )


def load_interface_records(path: Path = DEFAULT_MATRIX_PATH) -> list[InterfaceRecord]:
    """Reads the R1+R2 sheet only and returns its 203 InterfaceRecords —
    "Interface Matrix R2" is deliberately not read (see module
    docstring). Raises clearly on a structurally unexpected file —
    callers that need graceful degradation (the web app) should catch
    and fall back, the same pattern generate_workbook.py already uses
    for this same file."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if PRIMARY_SHEET not in wb.sheetnames:
            raise ValueError(f"Sheet {PRIMARY_SHEET!r} not found. Found: {wb.sheetnames}")
        header, rows = _sheet_rows(wb[PRIMARY_SHEET])
        return [_record_from_primary_row(header, r) for r in rows]
    finally:
        wb.close()


def build_interface_index(records: list[InterfaceRecord]) -> dict[str, list[InterfaceRecord]]:
    """Groups records by Datadog application name — the join key
    against alert data's APPLICATION_NAME. Records with no Datadog app
    (blank, or one of the NOT_DATADOG_VALUES placeholders) are skipped
    — they can never be a join candidate."""
    index: dict[str, list[InterfaceRecord]] = defaultdict(list)
    for rec in records:
        if rec.datadog_app:
            index[rec.datadog_app].append(rec)
    return dict(index)


def load_interface_index(path: Path = DEFAULT_MATRIX_PATH) -> dict[str, list[InterfaceRecord]] | None:
    """The graceful-degradation entry point other modules should call.
    Returns None — never raises — if the file isn't present or fails
    to parse, matching generate_workbook.load_interface_index's
    contract for the same underlying file."""
    if not path.exists():
        return None
    try:
        return build_interface_index(load_interface_records(path))
    except Exception:
        return None


# --------------------------------------------------------------------
# Matching alert rows against the interface catalog
# --------------------------------------------------------------------

_NON_ALNUM_RE = re.compile(r"[^A-Z0-9]")
_TOKEN_RE = re.compile(r"[A-Z][A-Z0-9_]{3,}")


def _normalize(text: str) -> str:
    return _NON_ALNUM_RE.sub("", text.upper())


_API_AFFIX_RE = re.compile(r"^API_")
_SRV_AFFIX_RE = re.compile(r"_SRV\w*$")


def _identifier_tokens(text: str) -> set[str]:
    """Pulls identifier-shaped substrings (SAP API/BAPI/service names)
    out of a line of text, e.g. "GET /API_PRODUCTION_ORDER_2_SRV/A_..."
    -> {"API_PRODUCTION_ORDER_2_SRV", "A_..."}. Matched against alert
    text after normalizing both sides the same way, so hyphen/
    underscore/casing differences don't matter.

    Each raw token also contributes an "API_"/"_SRV"-stripped core
    form. This isn't optional polish: Datadog's own alert job-names
    are built from the interface's core action, not its literal SAP
    service name — e.g. an alert's "OUTBOUND_DELIVERY-POST-GOODS-ISSUE"
    corresponds to the matrix's "API_OUTBOUND_DELIVERY_SRV", but never
    contains the literal "API"/"SRV" substrings itself. Matching only
    the raw token missed this entirely (verified against real data —
    it silently misattributed the single largest real alert category
    to the wrong interface). Both forms are kept since some alert text
    does carry the full "API_..._SRV" name verbatim (e.g. inside a
    literal URL), so stripping unconditionally would lose those hits.
    """
    tokens = set()
    for t in _TOKEN_RE.findall(text.upper()):
        if len(t) < 4:
            continue
        tokens.add(_normalize(t))
        core = _SRV_AFFIX_RE.sub("", _API_AFFIX_RE.sub("", t))
        if len(core) >= 4:
            tokens.add(_normalize(core))
    return tokens


def _alert_haystack(row: dict) -> str:
    parts = [
        str(row.get("ALERT_MESSAGE") or ""),
        str(row.get("ERROR_MESSAGE") or ""),
        str(row.get("ERROR_DETAILS") or ""),
        str(row.get("TRANSACTION_URL") or ""),
        str(row.get("BUSINESS_OBJECT_KEY") or ""),
    ]
    return _normalize(" ".join(parts))


@dataclass
class ScoreDetail:
    strong: int = 0
    weak: int = 0
    reasons: list[str] = field(default_factory=list)


def score_candidate(record: InterfaceRecord, haystack: str) -> ScoreDetail:
    detail = ScoreDetail()

    if record.datadog_business_object:
        dbo = _normalize(record.datadog_business_object)
        if dbo and dbo in haystack:
            detail.strong += 1
            detail.reasons.append(f"business_object:{record.datadog_business_object}")

    for api_id in record.api_identifiers:
        for tok in _identifier_tokens(api_id):
            if tok in haystack:
                detail.strong += 1
                detail.reasons.append(f"api:{tok}")
                break  # one strong hit per api_identifiers line is enough
        else:
            continue
        break  # one strong hit from this signal overall is enough to count it

    # Weak signal: interface name + free-text description, tokenized
    # the same way. The description matters, not just the name — see
    # the field's docstring on InterfaceRecord for the real case this
    # fixes (two candidates sharing one API identifier, only the
    # description text actually distinguishes which one an alert means).
    weak_text = " ".join(t for t in (record.interface_name, record.aif_description) if t)
    if weak_text:
        # set(), not a plain list: some matrix rows have their
        # description duplicated verbatim (a real data-quality quirk,
        # confirmed on ASTRO-WMS-BENE's "Outbound Delivery" row —
        # "Interface Outbound Delivery to WMS Astro" appears twice).
        # Counting every repetition let that row out-score the actual
        # correct match by sheer token repetition rather than genuine
        # relevance — each distinct word should count once per candidate.
        name_tokens = {t.upper() for t in re.split(r"[^A-Za-z0-9]+", weak_text) if len(t) >= 4}
        detail.weak = sum(1 for t in name_tokens if t in haystack)

    return detail


# The pseudo-stream a caller should group an unmatched/unavailable
# alert under, so it's still visible somewhere (the old page's fake
# "Intransit from SAP, MULESOFT" channel) rather than silently dropped.
# MatchResult.stream is None for those rows; callers substitute this.
STREAM_UNMAPPED = "Unmapped"


@dataclass
class MatchResult:
    state: str  # "exact" | "inferred" | "ambiguous" | "unmatched"
    record: InterfaceRecord | None
    stream: str | None
    also_streams: list[str]
    unmapped_app: str | None
    score: ScoreDetail | None = None


def _deterministic_pick(candidates: list[InterfaceRecord]) -> InterfaceRecord:
    """A stable, reproducible tie-break — same input always gives the
    same output, which is what keeps reconciliation possible even for
    the ambiguous case. Sorted by interface name then stream."""
    return sorted(candidates, key=lambda c: (c.interface_name or "", c.primary_stream() or ""))[0]


def resolve_alert_interface(app_name: str | None, index: dict[str, list[InterfaceRecord]], row: dict) -> MatchResult:
    if not app_name:
        return MatchResult("unmatched", None, None, [], None)

    candidates = index.get(app_name, [])
    if not candidates:
        return MatchResult("unmatched", None, None, [], app_name)

    def as_result(state: str, rec: InterfaceRecord, score: ScoreDetail | None) -> MatchResult:
        streams = rec.streams or []
        return MatchResult(state, rec, streams[0] if streams else None, streams[1:], None, score)

    if len(candidates) == 1:
        return as_result("exact", candidates[0], None)

    haystack = _alert_haystack(row)
    scored = [(score_candidate(c, haystack), c) for c in candidates]
    strong = [(s, c) for s, c in scored if s.strong >= 1]

    if len(strong) == 1:
        return as_result("exact", strong[0][1], strong[0][0])

    if len(strong) > 1:
        ranked = sorted(strong, key=lambda sc: (-sc[0].strong, -sc[0].weak))
        top, second = ranked[0], ranked[1]
        if (top[0].strong, top[0].weak) > (second[0].strong, second[0].weak):
            return as_result("inferred", top[1], top[0])
        return as_result("ambiguous", _deterministic_pick(candidates), top[0])

    # No strong signal from anything — fall back to weak (interface-name
    # token overlap) as a lower-confidence "inferred" tier.
    ranked_weak = sorted(scored, key=lambda sc: -sc[0].weak)
    if ranked_weak[0][0].weak > 0 and (len(ranked_weak) == 1 or ranked_weak[0][0].weak > ranked_weak[1][0].weak):
        return as_result("inferred", ranked_weak[0][1], ranked_weak[0][0])

    return as_result("ambiguous", _deterministic_pick(candidates), None)


# --------------------------------------------------------------------
# Self-check CLI
# --------------------------------------------------------------------

def _self_check(matrix_path: Path, alert_file: Path) -> None:
    import generate_workbook as gw

    print(f"Matrix: {matrix_path}")
    records = load_interface_records(matrix_path)
    print(f"  {len(records)} interface records loaded "
          f"({sum(1 for r in records if r.datadog_app)} joinable, "
          f"{sum(1 for r in records if r.not_datadog_monitored)} explicitly not-Datadog-monitored)")
    index = build_interface_index(records)
    print(f"  {len(index)} distinct joinable Datadog application names")
    print()

    print(f"Alert file: {alert_file}")
    rows = gw.read_generated_rows(alert_file)
    print(f"  {len(rows)} alert rows")
    print()

    state_counts = Counter()
    unmatched_apps = Counter()
    for r in rows:
        result = resolve_alert_interface(r.get("APPLICATION_NAME"), index, r)
        state_counts[result.state] += 1
        if result.state == "unmatched" and result.unmapped_app:
            unmatched_apps[result.unmapped_app] += 1

    print("Match-state histogram:")
    for state in ("exact", "inferred", "ambiguous", "unmatched"):
        n = state_counts.get(state, 0)
        pct = (n / len(rows) * 100) if rows else 0
        print(f"  {state:10s} {n:5d}  ({pct:5.1f}%)")
    print()

    print("Top unmatched application names (by alert count):")
    for app, count in unmatched_apps.most_common(20):
        print(f"  {count:5d}  {app!r}")
    if not unmatched_apps:
        print("  (none)")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--self-check", metavar="ALERT_XLSX", help="Run the match-state histogram against a generated alert workbook")
    parser.add_argument("--matrix", default=str(DEFAULT_MATRIX_PATH), help="Path to the interface matrix workbook")
    args = parser.parse_args()

    if args.self_check:
        _self_check(Path(args.matrix), Path(args.self_check))
    else:
        parser.print_help()


if __name__ == "__main__":
    main()

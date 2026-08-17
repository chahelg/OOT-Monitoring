"""
Alert Monitoring Workbook Generator.

Merges a daily Technical export and a daily Functional export into the
"Data" sheet of a copy of the reference/template workbook, wiring up the
same live formulas (Timestamp Readable / Alert type / Check) and leaving
the Match sheet and pivot table (Sheet2) untouched, marked to refresh the
moment the file is opened in Excel.

Accepts .csv or .xlsx source files (Datadog exports both; headers are
matched by name, not position).

Usage:
    py generate_workbook.py --technical "technical-active-alert.csv" --functional "functional-active-alert.csv" ^
        --template "5th Aug.xlsx" --output "output\active-alert category 5 Aug.xlsx"
"""

import argparse
import concurrent.futures
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import urllib.error
import urllib.request
import zipfile
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

APP_DIR = Path(__file__).parent
OUTPUT_DIR = APP_DIR / "output"

EXPECTED_HEADERS = [
    "TIMESTAMP",
    "SERVICE",
    "APPLICATION_NAME",
    "BUSINESS_OBJECT_KEY",
    "SEQUENCE",
    "TRANSACTION_URL",
    "ALERT_MESSAGE",
    "ERROR_MESSAGE",
    "ERROR_DETAILS",
]

# Output columns E..M map 1:1 to EXPECTED_HEADERS, in order.
DATA_HEADERS = [
    "alert grouping ",
    "Timestamp Readable ",
    "Alert type",
    "Check",
] + EXPECTED_HEADERS

_ILLEGAL_XML_RE = re.compile(
    "[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]"
)


def sanitize_xml_text(value: str) -> str:
    return _ILLEGAL_XML_RE.sub("", value)


def xml_escape(value: str) -> str:
    value = sanitize_xml_text(value)
    return (
        value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


_NUMERIC_STRING_RE = re.compile(r"-?\d+")


def coerce_csv_value(value):
    """CSV cells arrive as plain strings; numbers that look like whole
    numbers (timestamps, sequence numbers, purely-numeric business object
    keys) are converted to int so they land in the sheet as real numbers,
    exactly like an xlsx source would report them.

    Only the *emptiness* and *numeric-ness* checks look at a stripped
    copy — the returned value for real text is the original, untouched.
    Free-text fields (ALERT_MESSAGE etc.) can carry meaningful leading/
    trailing spaces baked into the source log line, and the spec requires
    columns E-M to be copied unchanged; stripping every field here was
    silently rewriting that text.
    """
    if value is None:
        return None
    stripped = value.strip()
    if stripped == "":
        return None
    if _NUMERIC_STRING_RE.fullmatch(stripped):
        return int(stripped)
    return value


def _rows_to_records(header_row, rows_iter, label: str, path_name: str, coerce) -> list[dict]:
    header_map = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        header_map[str(cell).strip().upper()] = idx

    missing = [h for h in EXPECTED_HEADERS if h not in header_map]
    if missing:
        raise ValueError(
            f"{path_name}: missing expected column header(s) {missing}. "
            f"Found headers: {list(header_map.keys())}"
        )

    rows = []
    for raw_row in rows_iter:
        if raw_row is None or all(v is None or v == "" for v in raw_row):
            continue
        raw_timestamp = raw_row[header_map["TIMESTAMP"]] if header_map["TIMESTAMP"] < len(raw_row) else None
        timestamp = coerce(raw_timestamp)
        if timestamp is None:
            continue
        record = {"grouping": label, "TIMESTAMP": int(timestamp)}
        for h in EXPECTED_HEADERS[1:]:
            idx = header_map[h]
            record[h] = coerce(raw_row[idx]) if idx < len(raw_row) else None
        rows.append(record)
    return rows


def read_source_rows(path: Path, label: str) -> list[dict]:
    if path.suffix.lower() == ".csv":
        with open(path, encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return []
        return _rows_to_records(rows[0], rows[1:], label, path.name, coerce_csv_value)

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        return _rows_to_records(header_row, rows_iter, label, path.name, lambda v: v)
    finally:
        wb.close()


def merge_rows(technical_rows: list[dict], functional_rows: list[dict]) -> list[dict]:
    # Matches the manual SOP: all Functional rows first (export order),
    # then all Technical rows below (export order) — not sorted/interleaved.
    return functional_rows + technical_rows


def load_match_rules(template_path: Path) -> list[tuple[str, str]]:
    """Reads Match!A2:B<last> (String a / String b) from the template so
    unmatched error text can be checked in Python, without needing Excel."""
    wb = openpyxl.load_workbook(template_path, data_only=True, read_only=True)
    try:
        ws = wb["Match"]
        rules = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            string_a, string_b = row[0], row[1]
            if string_a is None or string_b is None:
                continue
            rules.append((str(string_a), str(string_b)))
        return rules
    finally:
        wb.close()


def classify_text(error_message, error_details) -> str:
    """Mirrors column C's IF(M="null", use L, use M) branch — which text
    the Alert-type formula actually classifies against."""
    m = "" if error_details is None else str(error_details)
    if m == "null":
        return "" if error_message is None else str(error_message)
    return m


def find_unmatched(rows: list[dict], rules: list[tuple[str, str]]) -> list[tuple[str, int]]:
    """Rows whose classification text matches no Match rule — these are
    the ones that would show blank in column C / #N/A in column D, i.e.
    a new error type not yet noted in the Match sheet. Returns distinct
    texts with their occurrence count, most frequent first."""
    counts: dict[str, int] = {}
    for rec in rows:
        text = classify_text(rec.get("ERROR_MESSAGE"), rec.get("ERROR_DETAILS"))
        if not text:
            continue
        text_lower = text.lower()
        matched = any(a.lower() in text_lower and b.lower() in text_lower for a, b in rules)
        if not matched:
            counts[text] = counts.get(text, 0) + 1
    return sorted(counts.items(), key=lambda kv: kv[1], reverse=True)


def load_match_rules_full(template_path: Path) -> list[dict]:
    """Reads Match!A2:D<last> (String a / String b / Result / check) in
    sheet order — full detail needed for classification, for rendering
    the editable Match table, and for rewriting it."""
    wb = openpyxl.load_workbook(template_path, data_only=True, read_only=True)
    try:
        ws = wb["Match"]
        rules = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            string_a = row[0] if len(row) > 0 else None
            string_b = row[1] if len(row) > 1 else None
            result = row[2] if len(row) > 2 else None
            check = row[3] if len(row) > 3 else None
            if string_a is None or string_b is None:
                continue
            rules.append({
                "string_a": str(string_a),
                "string_b": str(string_b),
                "result": "" if result is None else str(result),
                "check": 1 if check == 1 else None,
            })
        return rules
    finally:
        wb.close()


def compute_alert_type(text: str, rules_full: list[dict]) -> str:
    """Mirrors column C's TEXTJOIN(", ", TRUE, FILTER(Result,
    ISNUMBER(SEARCH(a,text)) * ISNUMBER(SEARCH(b,text)))) — every rule
    whose String a AND String b both appear in `text` contributes its
    Result, joined in sheet order."""
    if not text:
        return ""
    text_lower = text.lower()
    matches = []
    for rule in rules_full:
        if rule["string_a"].lower() in text_lower and rule["string_b"].lower() in text_lower:
            if rule["result"]:
                matches.append(rule["result"])
    return ", ".join(matches)


def compute_check(alert_type: str, rules_full: list[dict]):
    """Mirrors VLOOKUP(alert_type, Match!C:D, 2, 0): first rule (in sheet
    order) whose Result exactly equals alert_type. Blank check -> 0
    (VLOOKUP's own behavior for a blank cell). No match -> the "#N/A"
    sentinel, same error Excel would show."""
    for rule in rules_full:
        if rule["result"] == alert_type:
            return rule["check"] if rule["check"] is not None else 0
    return "#N/A"


def epoch_ms_to_datetime(ms) -> datetime:
    """Mirrors column B's formula exactly: naive UTC arithmetic (whole
    seconds from the first 10 digits, milliseconds from the last 3,
    divided down to days, offset from 1970-01-01) — no timezone
    conversion, matching how the Excel formula behaves."""
    return datetime(1970, 1, 1) + timedelta(milliseconds=int(ms))


def compute_grouping(source_label: str, check) -> str:
    """Mirrors column A's IF(IFERROR(D,0)=1,"Functional",label)."""
    return "Functional" if check == 1 else source_label


_GROUPING_FORMULA_RE = re.compile(r'"Functional","([^"]*)"\)')


def read_generated_rows(xlsx_path: Path) -> list[dict]:
    """Reads an already-generated workbook's Data sheet and recomputes
    what Excel's live formulas would show, entirely in Python — no Excel
    required. Powers the in-browser Data/Report views. Column A is read
    as a formula string (not a cached value, since there is none) and its
    baked-in source-label fallback is pulled back out with regex."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=False, read_only=True)
    try:
        ws = wb["Data"]
        raw_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            e = row[4] if len(row) > 4 else None
            if e is None:
                continue
            a_formula = row[0] if len(row) > 0 else None
            m = _GROUPING_FORMULA_RE.search(str(a_formula) if a_formula else "")
            raw_rows.append({
                "source_label": m.group(1) if m else "",
                "TIMESTAMP": e,
                "SERVICE": row[5],
                "APPLICATION_NAME": row[6],
                "BUSINESS_OBJECT_KEY": row[7],
                "SEQUENCE": row[8],
                "TRANSACTION_URL": row[9],
                "ALERT_MESSAGE": row[10],
                "ERROR_MESSAGE": row[11],
                "ERROR_DETAILS": row[12],
            })
    finally:
        wb.close()

    rules_full = load_match_rules_full(xlsx_path)

    results = []
    for rec in raw_rows:
        text = classify_text(rec["ERROR_MESSAGE"], rec["ERROR_DETAILS"])
        alert_type = compute_alert_type(text, rules_full)
        check = compute_check(alert_type, rules_full)
        grouping = compute_grouping(rec["source_label"], check if isinstance(check, int) else None)
        results.append({
            "grouping": grouping,
            "timestamp_readable": epoch_ms_to_datetime(rec["TIMESTAMP"]),
            "alert_type": alert_type,
            "check": check,
            **{k: rec[k] for k in (
                "TIMESTAMP", "SERVICE", "APPLICATION_NAME", "BUSINESS_OBJECT_KEY",
                "SEQUENCE", "TRANSACTION_URL", "ALERT_MESSAGE", "ERROR_MESSAGE", "ERROR_DETAILS",
            )},
        })
    return results


def compute_pivot(rows: list[dict]) -> dict:
    """Reproduces the real pivot's shape for browser viewing: nested rows
    (alert grouping -> Alert type), columns = distinct days ascending,
    values = row count, with row/column totals. This is a read-only
    reproduction for convenience — the real pivot table Excel refreshes
    on open remains authoritative for anything sent externally."""
    cells: dict = defaultdict(lambda: defaultdict(int))
    row_keys = []
    seen = set()
    days_seen = set()

    for rec in rows:
        key = (rec["grouping"] or "", rec["alert_type"] or "")
        if key not in seen:
            seen.add(key)
            row_keys.append(key)
        day = rec["timestamp_readable"].strftime("%Y-%m-%d")
        cells[key][day] += 1
        days_seen.add(day)

    row_keys.sort()
    days = sorted(days_seen)
    row_totals = {k: sum(cells[k].values()) for k in row_keys}
    col_totals = {d: sum(cells[k].get(d, 0) for k in row_keys) for d in days}

    return {
        "row_keys": row_keys,
        "days": days,
        "cells": cells,
        "row_totals": row_totals,
        "col_totals": col_totals,
        "grand_total": sum(row_totals.values()),
    }


def compute_aging(rows: list[dict], now: datetime | None = None, threshold_hours: int = 48) -> list[dict]:
    """Groups rows by Alert type (the same categorization the daily
    observations email uses) and flags categories whose earliest
    occurrence is older than `threshold_hours` — i.e. still open (still
    present in this file) past the expected resolution window. Powers
    both the Aging tab and the Email Draft tab.

    Also flags a distinct "new spike" case: a young (not yet flagged)
    category where almost all its volume landed on the single most
    recent day in the dataset — these two flags are mutually exclusive
    by construction (a category already old enough to be `flagged` is
    reported there, not as a spike, even if its volume is recent).
    """
    if now is None:
        now = datetime.now()

    groups: dict[str, dict] = {}
    latest_overall_date = None
    for rec in rows:
        key = rec["alert_type"] or ""
        if not key:
            continue
        ts = rec["timestamp_readable"]
        day = ts.date()
        if latest_overall_date is None or day > latest_overall_date:
            latest_overall_date = day
        g = groups.setdefault(key, {
            "alert_type": key,
            "groupings": defaultdict(int),
            "day_counts": defaultdict(int),
            "count": 0,
            "earliest": ts,
            "latest": ts,
        })
        g["count"] += 1
        g["groupings"][rec["grouping"] or ""] += 1
        g["day_counts"][day] += 1
        if ts < g["earliest"]:
            g["earliest"] = ts
        if ts > g["latest"]:
            g["latest"] = ts

    results = []
    for g in groups.values():
        age_hours = (now - g["earliest"]).total_seconds() / 3600
        grouping = max(g["groupings"].items(), key=lambda kv: kv[1])[0]
        day_counts = sorted(g["day_counts"].items())
        flagged = age_hours > threshold_hours

        recent_cutoff = latest_overall_date - timedelta(days=1) if latest_overall_date else None
        is_recurring = (
            len(day_counts) >= 2
            and recent_cutoff is not None
            and any(d >= recent_cutoff for d, _ in day_counts)
        )

        latest_day_count = g["day_counts"].get(latest_overall_date, 0) if latest_overall_date else 0
        is_new_spike = (
            not flagged
            and latest_overall_date is not None
            and g["latest"].date() == latest_overall_date
            and g["count"] > 0
            and latest_day_count >= 0.8 * g["count"]
        )

        results.append({
            "alert_type": g["alert_type"],
            "grouping": grouping,
            "count": g["count"],
            "earliest": g["earliest"],
            "latest": g["latest"],
            "age_hours": age_hours,
            "age_days": int(age_hours // 24),
            "flagged": flagged,
            "day_counts": day_counts,  # list of (date, count), ascending
            "is_recurring": is_recurring,
            "is_new_spike": is_new_spike,
        })

    flagged_results = sorted((r for r in results if r["flagged"]), key=lambda r: -r["age_hours"])
    other_results = sorted((r for r in results if not r["flagged"]), key=lambda r: -r["count"])
    return flagged_results + other_results


def compute_aging_stats(aging: list[dict], top_n: int = 10) -> dict:
    """Summary numbers + a magnitude-ranked top-N slice for the Aging tab
    dashboard (stat tiles + bar chart) — kept separate from `compute_aging`
    itself since the table (worst-age-first) and the chart (ranked by
    count) have different jobs and sort orders."""
    flagged = [r for r in aging if r["flagged"]]
    spikes = [r for r in aging if r["is_new_spike"]]
    top = sorted(aging, key=lambda r: -r["count"])[:top_n]
    return {
        "n_flagged": len(flagged),
        "n_spikes": len(spikes),
        "flagged_total": sum(r["count"] for r in flagged),
        "oldest_days": max((r["age_days"] for r in flagged), default=0),
        "max_count": max((r["count"] for r in aging), default=1),
        "top": top,
    }


def build_email_draft(rows: list[dict], aging: list[dict], threshold_hours: int = 48, max_items: int = 6) -> str:
    """Plain-text daily observations email draft in the style of the
    team's existing daily email: a numbered paragraph per aging category
    (count, day-by-day breakdown, age — all facts) plus one disclosed,
    objective rule for the recurring-vs-new-spike framing sentence (see
    compute_aging's docstring for the exact rule). This is a draft to
    review before sending, not a final email — this tool has no
    send capability at all, only text generation.

    On a file where the rolling-window data means most/all categories
    are past the 48h threshold, `flagged` can be large (seen: all 18 of
    18) — writing an individual paragraph for every one of them isn't a
    usable email. Only the highest-volume `max_items` get a full
    write-up; the rest are named in one rollup line instead of silently
    dropped.
    """
    lines = ["Good morning Team,", "Sharing the daily alert summary.", "", "Observations:"]

    flagged_all = sorted((r for r in aging if r["flagged"]), key=lambda r: -r["count"])
    flagged = flagged_all[:max_items]
    overflow = flagged_all[max_items:]
    spikes = [r for r in aging if r["is_new_spike"]]

    n = 0
    for r in flagged:
        n += 1
        count = r["count"]
        there_are = "There is" if count == 1 else "There are"
        issue_word = "issue" if count == 1 else "issues"
        lines.append(f'{n}.\t{there_are} {count} {issue_word} in the "{r["alert_type"]}" category. '
                      f'These issues occurred on: ')
        for d, c in r["day_counts"]:
            lines.append(f'\to\t{d.strftime("%d-%m-%Y")} – {c} alert{"s" if c != 1 else ""}')
        days_past = r["age_days"] - (threshold_hours // 24)
        remain_word = "remains" if count == 1 else "remain"
        sentence = (
            f'All {count} {remain_word} open - these are now {r["age_days"]} days old '
            f'- {days_past} days past the 48-hour expected resolution window'
        )
        if r["is_recurring"]:
            sentence += (
                ', with new alerts continuing to appear across multiple days rather than '
                'being closed out, indicating this is a recurring issue rather than a one-time spike.'
            )
        else:
            sentence += '.'
        lines.append(sentence)
        lines.append("")

    for r in spikes:
        n += 1
        latest_day, latest_day_count = r["day_counts"][-1] if r["day_counts"] else (None, r["count"])
        alert_word = "alert" if latest_day_count == 1 else "alerts"
        prior = r["count"] - latest_day_count
        prior_clause = f", up from just {prior} previously" if prior > 0 else ""
        lines.append(
            f'{n}.\tA new spike is observed in the "{r["alert_type"]}" category, with '
            f'{latest_day_count} {alert_word} on {latest_day.strftime("%d-%m-%Y") if latest_day else "the latest day"} '
            f'alone{prior_clause} - this is a fresh issue and not a carryover from the earlier backlog.'
        )
        lines.append("")

    if overflow:
        n += 1
        names = ", ".join(f'"{r["alert_type"]}"' for r in overflow[:5])
        more = f", and {len(overflow) - 5} more" if len(overflow) > 5 else ""
        lines.append(
            f'{n}.\t{len(overflow)} more categories are also past the 48-hour window but at lower '
            f'volume: {names}{more}.'
        )
        lines.append("")

    technical_total = sum(1 for r in rows if r["grouping"] == "Technical")
    functional_total = sum(1 for r in rows if r["grouping"] == "Functional")
    latest_date = max((r["timestamp_readable"].date() for r in rows), default=None)
    latest_date_total = sum(1 for r in rows if r["timestamp_readable"].date() == latest_date) if latest_date else 0

    if flagged or spikes:
        n += 1
        if technical_total == functional_total:
            split_clause = f'technical and functional issues are evenly split ({technical_total} each)'
        else:
            bigger_label, bigger, smaller_label, smaller = (
                ("Technical", technical_total, "functional", functional_total)
                if technical_total > functional_total
                else ("Functional", functional_total, "technical", technical_total)
            )
            split_clause = f'{bigger_label} issues continue to dominate over {smaller_label} issues ({bigger} vs {smaller})'
        tail = (
            f', with {latest_date.strftime("%d-%m-%Y")} accounting for '
            f'{latest_date_total} alert{"s" if latest_date_total != 1 else ""} that day.'
            if latest_date else '.'
        )
        lines.append(f'{n}.\tOverall, {split_clause}' + tail)
        lines.append("")

    if flagged:
        top_names = ", ".join(f'"{r["alert_type"]}"' for r in flagged[:3])
        n += 1
        lines.append(
            f'{n}.\tIssues from previous days continue to remain unresolved in the system, '
            f'particularly around {top_names}, suggesting the expected {threshold_hours}-hour '
            f'turnaround time is not being consistently met.'
        )

    return "\n".join(lines).rstrip() + "\n"


# Local-only by design: this calls Ollama on 127.0.0.1, never an external
# API — the alert data (error text, business object keys, service names,
# internal URLs) never leaves this machine. See build_email_draft_ai.
OLLAMA_URL = "http://localhost:11434"
OLLAMA_EMAIL_MODEL = "qwen2.5:7b"


def _ordinal_day(n: int) -> str:
    suffix = "th" if 11 <= n % 100 <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def _select_notable_categories(aging: list[dict], max_items: int = 6) -> list[dict]:
    """Which categories earn an individual write-up in the AI draft.
    Deliberately broader than build_email_draft()'s fixed flagged/spike
    rule — a real analyst also calls out things like a high-volume
    category or a sharp decline, neither of which is "flagged" or a
    "new spike" by that rule's definition (see the real example email
    this feature was modeled on: its 4th item was neither, just a
    volume drop worth a mention). Selection itself stays deterministic
    Python, not the LLM's call — only the analytical wording below is
    model-written, which is what keeps this reliable on a small local
    model (see build_email_draft_ai's docstring)."""
    max_count = max((r["count"] for r in aging), default=1)
    scored = []
    for r in aging:
        reasons = []
        if r["flagged"]:
            reasons.append(f"open {r['age_days']}d, past the 48h SLA")
        if r["is_new_spike"]:
            reasons.append("almost all volume landed on the single latest day")
        if r["is_recurring"]:
            reasons.append("alerts keep appearing across multiple days including recently")
        if r["count"] >= max(8, max_count * 0.15):
            reasons.append("notably high volume")
        if len(r["day_counts"]) >= 2:
            peak = max(c for _, c in r["day_counts"])
            latest = r["day_counts"][-1][1]
            if peak >= 10 and latest <= peak * 0.5:
                reasons.append(f"down sharply from a peak of {peak} to {latest} most recently")
        if reasons:
            scored.append((r, reasons))
    scored.sort(key=lambda item: -item[0]["count"])
    if not scored:
        scored = [(r, ["among today's highest-volume categories"]) for r in sorted(aging, key=lambda r: -r["count"])[:3]]
    return scored[:max_items]


def _describe_trend(day_counts: list[tuple]) -> str:
    """Labels a category's day-by-day shape as an unambiguous fact
    handed to the model, rather than something it has to work out from
    the raw numbers itself. Directly motivated by a demonstrated failure
    mode: even told to "ground it strictly in the numbers", the model
    would occasionally call a 2-to-1 drop "a rise" or invent a decline
    that wasn't in the data. Pre-computing the direction in Python and
    telling it not to contradict the label removes that failure mode the
    same way pre-computing every other fact already does."""
    counts = [c for _, c in day_counts]
    if len(counts) == 1:
        return f"SINGLE DAY ONLY — {counts[0]} alerts on one day, no other day to compare to. Do not claim any trend or comparison."
    first, last, peak = counts[0], counts[-1], max(counts)
    if last >= peak:
        return f"RISING — from {first} up to {last} (its highest point) across {len(counts)} days."
    if last <= min(counts) and last < first:
        return f"DECLINING — from {first} down to {last} across {len(counts)} days."
    if last < peak * 0.6:
        return f"PEAKED THEN DECLINED — peaked at {peak}, now at {last}, across {len(counts)} days."
    if last > first:
        return f"RISING OVERALL — from {first} to {last} across {len(counts)} days (not monotonic)."
    if last < first:
        return f"DECLINING OVERALL — from {first} down to {last} across {len(counts)} days (not monotonic)."
    return f"FLAT — steady around {last} across {len(counts)} days."


def _call_ollama(prompt: str, model: str, ollama_url: str, timeout: int) -> str:
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "stream": False,
        "options": {"temperature": 0.4},
    }
    req = urllib.request.Request(
        f"{ollama_url.rstrip('/')}/api/chat",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            body = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        try:
            msg = json.loads(e.read().decode("utf-8")).get("error", str(e))
        except Exception:
            msg = str(e)
        raise RuntimeError(f"Ollama error: {msg} — is model {model!r} pulled? Run `ollama pull {model}`.") from e
    except (urllib.error.URLError, TimeoutError) as e:
        raise RuntimeError(
            f"Couldn't reach Ollama at {ollama_url} (or it timed out) — is it running? "
            f"Start it with `ollama serve`, or open the Ollama app, then try again. ({e})"
        ) from e

    text = body.get("message", {}).get("content", "").strip()
    if not text:
        raise RuntimeError("Ollama returned an empty response — try again.")
    return text


def _strip_wrapping_quotes(text: str) -> str:
    """Strips a quote pair only if it truly wraps the whole string (both
    ends). Models sometimes quote just a category name mid-sentence
    ("X" is the priority because...) rather than the whole reply —
    str.strip('"') would wrongly eat that lone leading quote and leave
    the unmatched closing one dangling, so this checks both ends."""
    text = text.strip()
    if len(text) >= 2 and text[0] == '"' and text[-1] == '"':
        return text[1:-1].strip()
    return text


_LITERAL_DATE_RE = re.compile(r'\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b')


def _strip_literal_dates(text: str) -> str:
    """Defense-in-depth on top of the prompt's "never write a literal
    date" instruction: the model doesn't always follow that, and has
    gotten a date wrong when it didn't (attributing one category's count
    to the wrong day). Rather than trust instruction-following alone,
    this removes any DD-MM-YYYY-shaped token — with a leading
    preposition if present — so a wrong date can never reach the final
    email regardless of what the model did."""
    text = re.sub(r'\s*\b(?:on|from|since|as of|starting)\s+' + _LITERAL_DATE_RE.pattern, '', text, flags=re.IGNORECASE)
    text = _LITERAL_DATE_RE.sub('', text)
    text = re.sub(r'\s{2,}', ' ', text)
    text = re.sub(r'\s+([.,;:])', r'\1', text)
    return text.strip()


def build_email_draft_ai(
    rows: list[dict],
    aging: list[dict],
    model: str = OLLAMA_EMAIL_MODEL,
    ollama_url: str = OLLAMA_URL,
    timeout: int = 600,
) -> str:
    """AI-assisted version of the daily observations email. Every fact —
    which categories are mentioned, their counts, their day-by-day
    breakdown, the totals — is computed in plain Python, exactly like
    build_email_draft(); a local Ollama model (nothing ever leaves this
    machine) is only asked to fill in the analytical commentary for each
    one (spiking / recurring / newly-appeared / declining, in its own
    words) plus the closing summary's framing sentence.

    This split is deliberate, not just simpler: an early version handed
    the whole email — facts included — to the model in one shot, seeded
    with a real past email as a style example, and it partly reproduced
    that example's own numbers and category names instead of using
    today's data. A later version still batched all categories into one
    prompt and the model occasionally borrowed one category's count for
    another's sentence. A 7B model run locally just isn't reliable
    enough at that scale of open-ended, multi-fact generation to trust
    with the numbers. What's below instead makes one small, isolated
    call per category — each only ever sees that one category's own
    figures, so it has nothing else to confuse them with — and stitches
    the results into a Python-built skeleton that owns every number.
    That still leaves room for the model to misjudge *emphasis* (which
    is fine — that's the point), which is why this is still a draft to
    review before sending, same as build_email_draft().
    """
    if not aging:
        raise ValueError("No data to draft an email from — generate or select a file with rows first.")

    today = date.today()
    all_dates = [d for r in aging for d, _ in r["day_counts"]]
    if not all_dates:
        raise ValueError("No dated alert rows to draft an email from.")
    span_days = (max(all_dates) - min(all_dates)).days + 1
    total_alerts = sum(r["count"] for r in aging)
    latest_date = max(all_dates)
    latest_date_total = sum(c for r in aging for d, c in r["day_counts"] if d == latest_date)

    selected = _select_notable_categories(aging)
    top_category = selected[0][0]["alert_type"]
    # Only worth mentioning to the model at all if it's a real share of
    # the day's volume — otherwise a 7B model tends to dutifully mention
    # it anyway and call one stray alert a "meaningful concentration".
    # Deciding this in Python (rather than asking the model to judge it)
    # is the same principle as the rest of this function: never hand the
    # model a judgment call Python can just make correctly beforehand.
    latest_day_meaningful = latest_date_total >= max(5, total_alerts * 0.12)

    # One isolated call per category — NOT one big call covering all of
    # them. An earlier version batched all items into a single prompt
    # and the model would occasionally borrow one category's number for
    # another's sentence (e.g. claiming an 11-alert category "peaked at
    # 105", which was actually a different category's figure). Giving
    # each call only the one category's own numbers makes that
    # particular failure structurally impossible, not just less likely
    # — there's nothing else in its context to confuse it with. Costs
    # more calls, but each is small and the model stays loaded between
    # them, so it isn't much slower in practice.
    per_item_analysis = []
    for r, reasons in selected:
        day_str = "; ".join(f"{d.strftime('%d-%m-%Y')}: {c} alert(s)" for d, c in r["day_counts"])
        trend = _describe_trend(r["day_counts"])
        item_prompt = f"""You are a data analyst writing ONE short note for an internal ops status email. Plain, direct, analytical tone — never marketing language, never an apology, never claiming something is fixed/closed/resolved unless the numbers say so.

This alert category: "{r["alert_type"]}"
Total: {r["count"]}
Day-by-day (for context only): {day_str}
Confirmed trend direction — treat this as ground truth, do not contradict it or re-derive your own direction from the day-by-day numbers: {trend}
Why it stood out: {"; ".join(reasons)}

Write 1-2 sentences of analysis consistent with the confirmed trend direction above (e.g. if it says DECLINING, your sentence must describe a decline, never a "spike" or "rise"; if SINGLE DAY ONLY, do not claim any trend or comparison to other days at all). Do not restate the category name or the total count — those are already shown elsewhere. Never write out a specific date yourself (they're already shown as bullets above you) — refer to timing only in relative terms if needed ("the most recent day", "earlier in the week"). Do not restate individual day counts from the day-by-day breakdown — describe the overall shape, not each value. Start straight into the analysis. Do not mention or compare to any other category, you have no information about any other category. Reply with ONLY the 1-2 sentences, nothing else — no preamble, no quotes, no label."""
        analysis = _strip_literal_dates(_strip_wrapping_quotes(_call_ollama(item_prompt, model, ollama_url, timeout)))
        per_item_analysis.append(analysis)

    totals_list = "\n".join(f'- "{r["alert_type"]}": {r["count"]} total' for r, _ in selected)
    summary_prompt = f"""You are a data analyst writing ONE closing sentence for an internal ops status email.

Today's notable alert categories and their totals:
{totals_list}

"{top_category}" is today's top priority. Write ONE sentence (not a paragraph) explaining briefly why, grounded strictly in the totals above (e.g. its volume relative to the others, or that it's the largest). Do not restate the total alert count or repeat the list. Reply with ONLY that one sentence, nothing else — no preamble, no quotes, no label."""
    priority_reason = _strip_wrapping_quotes(_call_ollama(summary_prompt, model, ollama_url, timeout))

    lines = ["Good morning Team,", "Sharing the daily alert summary.", "",
              f"PFA: the data sheet for today- {_ordinal_day(today.day)} {today.strftime('%b')}", "",
              "Observations:"]
    for i, ((r, _reasons), analysis) in enumerate(zip(selected, per_item_analysis)):
        count = r["count"]
        there_are = "There is" if count == 1 else "There are"
        issue_word = "issue" if count == 1 else "issues"
        lines.append(f'{i + 1}.\t{there_are} {count} {issue_word} in the "{r["alert_type"]}" category. '
                      f'These issues occurred on the following dates:')
        for d, c in r["day_counts"]:
            lines.append(f'•\t{d.strftime("%d-%m-%Y")} – {c} alert{"s" if c != 1 else ""}')
        if analysis:
            lines.append(analysis)
        lines.append("")

    summary_line = f"Summary: Total alert volume is {total_alerts}."
    if latest_day_meaningful:
        summary_line += (
            f" {latest_date.strftime('%d-%m-%Y')} was the highest-volume day with "
            f"{latest_date_total} alert{'s' if latest_date_total != 1 else ''}."
        )
    if priority_reason:
        summary_line += f" {priority_reason}"
    lines.append(summary_line)
    lines.append("")
    lines.append(
        f"Note: Since the log data only retains a {span_days}-day rolling window, open issues from "
        f"before {min(all_dates).strftime('%d-%m-%Y')} are no longer visible in this data and may still "
        "be unresolved in the source system. Recommend the team separately verify the status of older "
        "open issues, as they can't be confirmed as closed just because they've dropped out of this report."
    )

    return "\n".join(lines).rstrip() + "\n"


def cell_xml(col: str, row_num: int, value, style: str | None = None) -> str:
    s_attr = f' s="{style}"' if style else ""
    if value is None or value == "":
        return f'<c r="{col}{row_num}"{s_attr}/>'
    if isinstance(value, bool):
        text = "TRUE" if value else "FALSE"
        return (
            f'<c r="{col}{row_num}"{s_attr} t="inlineStr">'
            f"<is><t>{xml_escape(text)}</t></is></c>"
        )
    if isinstance(value, (int, float)):
        num = int(value) if isinstance(value, float) and value.is_integer() else value
        return f'<c r="{col}{row_num}"{s_attr}><v>{num}</v></c>'
    text = str(value)
    preserve = ' xml:space="preserve"' if text != text.strip() or "\n" in text else ""
    return (
        f'<c r="{col}{row_num}"{s_attr} t="inlineStr">'
        f"<is><t{preserve}>{xml_escape(text)}</t></is></c>"
    )


def formula_cell_a(row_num: int, grouping: str) -> str:
    # Automates the manual "filter Check=1, set alert grouping to
    # Functional" step: Match!D=1 means this Alert type is inherently a
    # Functional issue regardless of which feed it was logged under, so
    # it overrides the source-file-based label. IFERROR guards against a
    # brand-new/unmatched error type, where D is #N/A rather than 1 or
    # blank — that must fall back to the source label, not propagate the
    # error into the grouping column (and the pivot table).
    grouping_escaped = xml_escape(grouping)
    formula = f'IF(IFERROR(D{row_num},0)=1,"Functional","{grouping_escaped}")'
    return f'<c r="A{row_num}" t="str"><f>{formula}</f></c>'


def formula_cell_b(row_num: int) -> str:
    formula = (
        f'((((LEFT(E{row_num},10) &amp; "." &amp; RIGHT(E{row_num},3))/60)/60)/24)'
        f"+DATE(1970,1,1)"
    )
    return f'<c r="B{row_num}" s="5"><f>{formula}</f></c>'


def formula_cell_c(row_num: int) -> str:
    r = row_num
    # Uses Table1 structured references (the Match sheet's table object)
    # instead of a hardcoded Match!$A$2:$A$55-style range, so the formula
    # keeps working automatically as rows are added to Match over time.
    formula = (
        f'IF(M{r}="null",_xlfn.TEXTJOIN(", ",TRUE,\n'
        f"_xlfn._xlws.FILTER(Table1[Result ],\n"
        f"ISNUMBER(SEARCH(Table1[String a],L{r}))*\n"
        f"ISNUMBER(SEARCH(Table1[String b],L{r})),\n"
        f'"")),_xlfn.TEXTJOIN(", ",TRUE,\n'
        f"_xlfn._xlws.FILTER(Table1[Result ],\n"
        f"ISNUMBER(SEARCH(Table1[String a],M{r}))*\n"
        f"ISNUMBER(SEARCH(Table1[String b],M{r})),\n"
        f'"")))'
    )
    return f'<c r="C{r}" t="str" cm="1"><f t="array" ref="C{r}">{formula}</f></c>'


def formula_cell_d(row_num: int) -> str:
    return f'<c r="D{row_num}"><f>VLOOKUP(C{row_num},Match!C:D,2,0)</f></c>'


def build_data_sheet_xml(rows: list[dict]) -> str:
    last_row = len(rows) + 1
    header_cells = []
    for col, header in zip("ABCDEFGHIJKLM", DATA_HEADERS):
        style = ' s="4"' if col in "EFGHIJKLM" else ""
        header_cells.append(
            f'<c r="{col}1"{style} t="inlineStr"><is>'
            f'<t xml:space="preserve">{xml_escape(header)}</t></is></c>'
        )
    header_row_xml = f'<row r="1" spans="1:13">{"".join(header_cells)}</row>'

    row_xmls = [header_row_xml]
    for i, rec in enumerate(rows):
        r = i + 2
        cells = [
            formula_cell_a(r, rec["grouping"]),
            formula_cell_b(r),
            formula_cell_c(r),
            formula_cell_d(r),
            cell_xml("E", r, rec["TIMESTAMP"]),
            cell_xml("F", r, rec["SERVICE"]),
            cell_xml("G", r, rec["APPLICATION_NAME"]),
            cell_xml("H", r, rec["BUSINESS_OBJECT_KEY"]),
            cell_xml("I", r, rec["SEQUENCE"]),
            cell_xml("J", r, rec["TRANSACTION_URL"]),
            cell_xml("K", r, rec["ALERT_MESSAGE"]),
            cell_xml("L", r, rec["ERROR_MESSAGE"]),
            cell_xml("M", r, rec["ERROR_DETAILS"]),
        ]
        row_xmls.append(f'<row r="{r}" spans="1:13">{"".join(cells)}</row>')

    sheet_data = "".join(row_xmls)

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
        'mc:Ignorable="x14ac xr xr2 xr3" '
        'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac" '
        'xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision" '
        'xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2" '
        'xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3" '
        'xr:uid="{48567247-C74A-4A61-8200-2F37C4F4E905}">'
        f'<dimension ref="A1:M{last_row}"/>'
        '<sheetViews><sheetView tabSelected="1" workbookViewId="0">'
        '<selection activeCell="A1" sqref="A1"/></sheetView></sheetViews>'
        '<sheetFormatPr defaultColWidth="37.6328125" defaultRowHeight="14.5" x14ac:dyDescent="0.35"/>'
        '<cols><col min="1" max="1" width="13.453125" bestFit="1" customWidth="1"/>'
        '<col min="3" max="3" width="65.54296875" customWidth="1"/>'
        '<col min="4" max="4" width="9.36328125" bestFit="1" customWidth="1"/>'
        '<col min="5" max="5" width="30.453125" customWidth="1"/>'
        '<col min="6" max="6" width="34.6328125" bestFit="1" customWidth="1"/>'
        '<col min="7" max="7" width="21.54296875" bestFit="1" customWidth="1"/>'
        '<col min="8" max="8" width="20.453125" bestFit="1" customWidth="1"/>'
        '<col min="9" max="9" width="10.08984375" bestFit="1" customWidth="1"/>'
        '<col min="10" max="10" width="121.6328125" bestFit="1" customWidth="1"/>'
        '<col min="11" max="11" width="196.36328125" bestFit="1" customWidth="1"/>'
        '<col min="12" max="12" width="64.6328125" customWidth="1"/></cols>'
        f"<sheetData>{sheet_data}</sheetData>"
        f'<autoFilter ref="A1:M{last_row}"/>'
        '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/>'
        '<pageSetup paperSize="9" orientation="portrait"/>'
        "</worksheet>"
    )


def patch_workbook_xml(xml_text: str) -> str:
    xml_text = re.sub(r'<sheet name="Detail1"[^/]*/>', "", xml_text)
    # Excel's stashed AutoFilter memory for the Data sheet — not required,
    # and its row count/localSheetId are stale template artifacts, so drop
    # it outright rather than trying to keep it in sync.
    xml_text = re.sub(
        r'<definedName name="_xlnm\._FilterDatabase"[^>]*>.*?</definedName>',
        "",
        xml_text,
    )
    xml_text = re.sub(
        r'<calcPr calcId="(\d+)"/>',
        r'<calcPr calcId="\1" fullCalcOnLoad="1"/>',
        xml_text,
    )
    xml_text = re.sub(r'activeTab="\d+"', 'activeTab="1"', xml_text)
    return xml_text


def patch_workbook_rels(xml_text: str, detail_rid: str, calc_rid: str | None) -> str:
    # detail_rid/calc_rid are resolved per-template (see generate_workbook)
    # rather than assumed to always be "rId1"/"rId10" — which sheet/part
    # physically lands at a given rId isn't stable across templates.
    xml_text = re.sub(
        rf'<Relationship Id="{detail_rid}" Type="[^"]*worksheet" Target="[^"]*"/>',
        "",
        xml_text,
    )
    if calc_rid:
        xml_text = re.sub(
            rf'<Relationship Id="{calc_rid}" Type="[^"]*calcChain" Target="calcChain\.xml"/>',
            "",
            xml_text,
        )
    return xml_text


def patch_content_types(xml_text: str, detail_sheet_part: str, detail_table_part: str | None) -> str:
    parts = ["/" + detail_sheet_part, "/xl/calcChain.xml"]
    if detail_table_part:
        parts.append("/" + detail_table_part)
    for part in parts:
        xml_text = re.sub(
            rf'<Override PartName="{re.escape(part)}"[^/]*/>', "", xml_text
        )
    return xml_text


def patch_app_xml(xml_text: str) -> str:
    xml_text = xml_text.replace(
        "<vt:variant><vt:i4>4</vt:i4></vt:variant>",
        "<vt:variant><vt:i4>3</vt:i4></vt:variant>",
    )
    xml_text = re.sub(r'(<vt:vector size=")4(" baseType="lpstr">)', r"\g<1>3\g<2>", xml_text)
    xml_text = xml_text.replace("<vt:lpstr>Detail1</vt:lpstr>", "")
    return xml_text


def patch_pivot_cache_definition(xml_text: str) -> str:
    if "refreshOnLoad" in xml_text:
        return xml_text
    return xml_text.replace(
        "<pivotCacheDefinition ",
        '<pivotCacheDefinition refreshOnLoad="1" ',
        1,
    )


def _sheet_rid(workbook_xml: str, sheet_name: str) -> str:
    """Finds a sheet's r:id by name (sheet order/IDs aren't guaranteed
    stable across templates — Excel names worksheet parts by creation
    history, not by tab position, so "sheet1.xml"/"rId1" can end up
    belonging to any tab depending on a template's own history)."""
    m = re.search(rf'<sheet[^>]*\bname="{re.escape(sheet_name)}"[^>]*/>', workbook_xml)
    if not m:
        raise ValueError(f"Sheet {sheet_name!r} not found in workbook.xml")
    rid_m = re.search(r'r:id="(rId\d+)"', m.group(0))
    if not rid_m:
        raise ValueError(f"No r:id found for sheet {sheet_name!r}")
    return rid_m.group(1)


def _rid_target(rels_xml: str, rid: str) -> str:
    target_m = re.search(rf'<Relationship Id="{rid}"[^>]*Target="([^"]+)"', rels_xml)
    if not target_m:
        raise ValueError(f"Relationship {rid} not found in workbook.xml.rels")
    return "xl/" + target_m.group(1).lstrip("/")


def _resolve_sheet_part(workbook_xml: str, rels_xml: str, sheet_name: str) -> str:
    """Finds which xl/worksheets/sheetN.xml a sheet name actually maps to
    (sheet order/IDs aren't guaranteed stable across templates)."""
    return _rid_target(rels_xml, _sheet_rid(workbook_xml, sheet_name))


def _resolve_related_part(zin: zipfile.ZipFile, sheet_part: str, type_suffix: str) -> str | None:
    """Finds a part attached to a worksheet via its own _rels file (e.g.
    its table, or a stashed customProperty bin), if any. Numeric suffixes
    on these parts (table1.xml vs table2.xml, customProperty1.bin vs
    customProperty2.bin) aren't stable across templates, so this always
    follows the relationship rather than assuming a number."""
    sheet_dir, sheet_file = sheet_part.rsplit("/", 1)
    rels_path = f"{sheet_dir}/_rels/{sheet_file}.rels"
    if rels_path not in zin.namelist():
        return None
    rels_xml = zin.read(rels_path).decode("utf-8")
    m = re.search(rf'Type="[^"]*?/{type_suffix}"\s+Target="([^"]+)"', rels_xml)
    if not m:
        return None
    parts = []
    for part in (Path(sheet_dir) / m.group(1)).as_posix().split("/"):
        if part == "..":
            parts.pop()
        else:
            parts.append(part)
    return "/".join(parts)


def _resolve_table_part(zin: zipfile.ZipFile, sheet_part: str) -> str | None:
    """Finds the table part (e.g. xl/tables/table2.xml) attached to a
    worksheet, if any, by following its own _rels file."""
    return _resolve_related_part(zin, sheet_part, "table")


def _match_rule_cell(col: str, row_num: int, text) -> str:
    text = "" if text is None else str(text)
    if text == "":
        return f'<c r="{col}{row_num}"/>'
    preserve = ' xml:space="preserve"' if text != text.strip() else ""
    return f'<c r="{col}{row_num}" t="inlineStr"><is><t{preserve}>{xml_escape(text)}</t></is></c>'


def add_match_rules(template_path: Path, new_rules: list[tuple[str, str, str, int | None]]) -> int:
    """Appends (String a, String b, Result, check) rows to a workbook's
    Match sheet in place, expanding Table1's range to include them —
    using the same zip-surgery approach as generate_workbook(), so the
    pivot table/cache and everything else is left completely untouched.

    Returns the new last row number.
    """
    tmp_path = template_path.with_name(template_path.name + f".tmp-{os.getpid()}")
    try:
        with zipfile.ZipFile(template_path, "r") as zin:
            names = zin.namelist()
            workbook_xml = zin.read("xl/workbook.xml").decode("utf-8")
            rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            sheet_part = _resolve_sheet_part(workbook_xml, rels_xml, "Match")
            match_xml = zin.read(sheet_part).decode("utf-8")

            last_row = max(int(m) for m in re.findall(r'<row r="(\d+)"', match_xml))

            new_rows_xml = []
            r = last_row
            for string_a, string_b, result, check in new_rules:
                r += 1
                cells = [
                    _match_rule_cell("A", r, string_a),
                    _match_rule_cell("B", r, string_b),
                    _match_rule_cell("C", r, result),
                ]
                if check:
                    cells.append(f'<c r="D{r}"><v>{int(check)}</v></c>')
                new_rows_xml.append(f'<row r="{r}" spans="1:4">{"".join(cells)}</row>')
            new_last_row = r

            match_xml = match_xml.replace("</sheetData>", "".join(new_rows_xml) + "</sheetData>")
            match_xml = re.sub(r"A1:D\d+", f"A1:D{new_last_row}", match_xml)

            table_part = _resolve_table_part(zin, sheet_part)
            table_xml = None
            if table_part:
                table_xml = re.sub(r"A1:D\d+", f"A1:D{new_last_row}", zin.read(table_part).decode("utf-8"))

            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for name in names:
                    data = zin.read(name)
                    if name == sheet_part:
                        data = match_xml.encode("utf-8")
                    elif table_part and name == table_part:
                        data = table_xml.encode("utf-8")
                    zout.writestr(name, data)
        os.replace(tmp_path, template_path)
    except BaseException:
        tmp_path.unlink(missing_ok=True)
        raise
    return new_last_row


def remove_last_match_rows(template_path: Path, count: int) -> int:
    """Removes the last `count` rule rows from a workbook's Match sheet
    in place (e.g. to undo/replace rows just added with a broader rule),
    shrinking Table1's range to match. Same zip-surgery approach as
    add_match_rules — pivot table/cache untouched. Returns the new last
    row number."""
    tmp_path = template_path.with_name(template_path.name + f".tmp-{os.getpid()}")
    try:
        with zipfile.ZipFile(template_path, "r") as zin:
            names = zin.namelist()
            workbook_xml = zin.read("xl/workbook.xml").decode("utf-8")
            rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            sheet_part = _resolve_sheet_part(workbook_xml, rels_xml, "Match")
            match_xml = zin.read(sheet_part).decode("utf-8")

            row_numbers = sorted(int(m) for m in re.findall(r'<row r="(\d+)"', match_xml))
            rows_to_drop = row_numbers[-count:]
            new_last_row = row_numbers[-count - 1]

            for r in rows_to_drop:
                match_xml = re.sub(rf'<row r="{r}"[^>]*>.*?</row>', "", match_xml)

            match_xml = re.sub(r"A1:D\d+", f"A1:D{new_last_row}", match_xml)

            table_part = _resolve_table_part(zin, sheet_part)
            table_xml = None
            if table_part:
                table_xml = re.sub(r"A1:D\d+", f"A1:D{new_last_row}", zin.read(table_part).decode("utf-8"))

            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for name in names:
                    data = zin.read(name)
                    if name == sheet_part:
                        data = match_xml.encode("utf-8")
                    elif table_part and name == table_part:
                        data = table_xml.encode("utf-8")
                    zout.writestr(name, data)
        os.replace(tmp_path, template_path)
    except BaseException:
        tmp_path.unlink(missing_ok=True)
        raise
    return new_last_row


def set_match_rules(template_path: Path, rules: list[dict]) -> int:
    """Rewrites the Match sheet's entire rule range (rows 2..N) from the
    given ordered list of {string_a, string_b, result, check} dicts — a
    general add/edit/delete/reorder primitive, unlike add_match_rules
    (append-only) or remove_last_match_rows (trim-from-end only). This is
    what backs the browser's editable Match tab. Same zip-surgery
    approach as the rest of this module; pivot table/cache untouched.
    Returns the new last row number.
    """
    tmp_path = template_path.with_name(template_path.name + f".tmp-{os.getpid()}")
    try:
        with zipfile.ZipFile(template_path, "r") as zin:
            names = zin.namelist()
            workbook_xml = zin.read("xl/workbook.xml").decode("utf-8")
            rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            sheet_part = _resolve_sheet_part(workbook_xml, rels_xml, "Match")
            match_xml = zin.read(sheet_part).decode("utf-8")

            existing_rows = [int(m) for m in re.findall(r'<row r="(\d+)"', match_xml) if int(m) >= 2]
            for r in existing_rows:
                match_xml = re.sub(rf'<row r="{r}"[^>]*>.*?</row>', "", match_xml)

            new_rows_xml = []
            r = 1
            for rule in rules:
                r += 1
                cells = [
                    _match_rule_cell("A", r, rule.get("string_a")),
                    _match_rule_cell("B", r, rule.get("string_b")),
                    _match_rule_cell("C", r, rule.get("result")),
                ]
                if rule.get("check"):
                    cells.append(f'<c r="D{r}"><v>1</v></c>')
                new_rows_xml.append(f'<row r="{r}" spans="1:4">{"".join(cells)}</row>')
            new_last_row = max(r, 1)

            match_xml = match_xml.replace("</sheetData>", "".join(new_rows_xml) + "</sheetData>")
            match_xml = re.sub(r"A1:D\d+", f"A1:D{new_last_row}", match_xml)

            table_part = _resolve_table_part(zin, sheet_part)
            table_xml = None
            if table_part:
                table_xml = re.sub(r"A1:D\d+", f"A1:D{new_last_row}", zin.read(table_part).decode("utf-8"))

            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for name in names:
                    data = zin.read(name)
                    if name == sheet_part:
                        data = match_xml.encode("utf-8")
                    elif table_part and name == table_part:
                        data = table_xml.encode("utf-8")
                    zout.writestr(name, data)
        os.replace(tmp_path, template_path)
    except BaseException:
        tmp_path.unlink(missing_ok=True)
        raise
    return new_last_row


def generate_workbook(technical_path: Path, functional_path: Path, template_path: Path, output_path: Path) -> dict:
    technical_rows = read_source_rows(technical_path, "Technical")
    functional_rows = read_source_rows(functional_path, "Functional")
    rows = merge_rows(technical_rows, functional_rows)

    if not rows:
        raise ValueError("No data rows found in either input file.")

    match_rules = load_match_rules(template_path)
    unmatched = find_unmatched(rows, match_rules)

    data_sheet_xml = build_data_sheet_xml(rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Write to a sibling temp file and swap it into place at the end. This
    # keeps a same-path template==output run (e.g. re-running on top of
    # today's own file, or an auto-detected "latest template" that happens
    # to match the output name) from reading and truncating the same file
    # at once, and avoids leaving a half-written .xlsx behind on error.
    tmp_path = output_path.with_name(output_path.name + f".tmp-{os.getpid()}")

    try:
        with zipfile.ZipFile(template_path, "r") as zin:
            names = zin.namelist()
            workbook_xml = zin.read("xl/workbook.xml").decode("utf-8")
            rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")

            # Which physical sheetN.xml each logical tab maps to is NOT
            # stable across templates — Excel names worksheet parts by
            # creation history, not by current tab position, so a
            # template that's been through more Excel saves can easily
            # have "Data" living at sheet2.xml instead of sheet3.xml, or
            # "Detail1" (the disposable staging sheet dropped below) at
            # sheet2.xml instead of sheet1.xml. Resolving both by name
            # (the same way Match-sheet edits already do) instead of by
            # a hardcoded number is what keeps this from silently
            # writing fresh data into the wrong tab, or deleting the
            # wrong sheet's relationship, on a template with different
            # internal numbering.
            data_part = _resolve_sheet_part(workbook_xml, rels_xml, "Data")
            detail_rid = _sheet_rid(workbook_xml, "Detail1")
            detail_sheet_part = _rid_target(rels_xml, detail_rid)
            detail_dir, detail_file = detail_sheet_part.rsplit("/", 1)
            detail_sheet_rels_part = f"{detail_dir}/_rels/{detail_file}.rels"
            detail_table_part = _resolve_table_part(zin, detail_sheet_part)
            detail_custom_property_part = _resolve_related_part(zin, detail_sheet_part, "customProperty")
            calc_rid_m = re.search(r'<Relationship Id="(rId\d+)"[^>]*Target="calcChain\.xml"', rels_xml)
            calc_rid = calc_rid_m.group(1) if calc_rid_m else None

            skip = {detail_sheet_part, detail_sheet_rels_part, "xl/calcChain.xml"}
            if detail_table_part:
                skip.add(detail_table_part)
            if detail_custom_property_part:
                skip.add(detail_custom_property_part)

            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for name in names:
                    if name in skip:
                        continue
                    data = zin.read(name)
                    if name == data_part:
                        data = data_sheet_xml.encode("utf-8")
                    elif name == "xl/workbook.xml":
                        data = patch_workbook_xml(data.decode("utf-8")).encode("utf-8")
                    elif name == "xl/_rels/workbook.xml.rels":
                        data = patch_workbook_rels(data.decode("utf-8"), detail_rid, calc_rid).encode("utf-8")
                    elif name == "[Content_Types].xml":
                        data = patch_content_types(
                            data.decode("utf-8"), detail_sheet_part, detail_table_part
                        ).encode("utf-8")
                    elif name == "docProps/app.xml":
                        data = patch_app_xml(data.decode("utf-8")).encode("utf-8")
                    elif name == "xl/pivotCache/pivotCacheDefinition1.xml":
                        data = patch_pivot_cache_definition(data.decode("utf-8")).encode("utf-8")
                    zout.writestr(name, data)
        os.replace(tmp_path, output_path)
    except BaseException:
        tmp_path.unlink(missing_ok=True)
        raise

    print(f"Wrote {output_path} ({len(rows)} data rows: "
          f"{len(technical_rows)} technical + {len(functional_rows)} functional)")
    if unmatched:
        print(f"{len(unmatched)} distinct error message(s) don't match any Match sheet rule "
              f"(will show blank Alert type / #N/A Check):")
        for text, count in unmatched:
            preview = text if len(text) <= 160 else text[:157] + "..."
            print(f"  [{count}x] {preview}")

    return {
        "row_count": len(rows),
        "technical_count": len(technical_rows),
        "functional_count": len(functional_rows),
        "unmatched": unmatched,
    }


def default_output_name() -> str:
    today = date.today()
    return f"active-alert category {today.day} {today.strftime('%b')}.xlsx"


def find_latest_template() -> Path | None:
    """Most recently generated workbook in the canonical output folder —
    used as the default template so Match-sheet edits carry forward."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    candidates = sorted(
        OUTPUT_DIR.glob("*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


_TRANSIENT_COM_HRESULTS = {-2147418111, -2147417846}  # RPC_E_CALL_REJECTED, RPC_E_SERVERCALL_RETRYLATER


def _process_alive(pid: int) -> bool:
    result = subprocess.run(
        ["tasklist", "/FI", f"PID eq {pid}"],
        capture_output=True, text=True,
        creationflags=subprocess.CREATE_NO_WINDOW,
    )
    return str(pid) in result.stdout


def _force_kill(pid: int) -> None:
    subprocess.run(
        ["taskkill", "/PID", str(pid), "/F"],
        capture_output=True,
        creationflags=subprocess.CREATE_NO_WINDOW,
    )


def _run_in_excel(xlsx_path: Path, action, refresh_pivot: bool = True, attempts: int = 6, retry_delay: float = 1.5):
    """Entry point for every "open a throwaway copy of a workbook in real
    Excel and do something with Sheet2" operation. Runs the actual work
    (`_run_in_excel_impl`) in a brand-new, dedicated thread every single
    call — root-caused this as necessary, not optional: the web app's
    Flask server runs single-threaded, so every request after the first
    reuses the same OS thread's COM apartment, and repeated
    CoInitialize/CoUninitialize cycles on that one thread degrade it —
    confirmed directly: the *first* Excel COM call in a process/thread
    always succeeded in testing, and *every* subsequent one on that same
    thread then failed with RPC_E_CALL_REJECTED regardless of how long a
    delay was inserted before retrying (ruling out a simple timing/
    contention issue). A fresh thread has never touched COM before, so
    it reliably reproduces the "always works" first-call state on every
    invocation, independent of whatever thread the caller runs on.
    """
    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(_run_in_excel_impl, xlsx_path, action, refresh_pivot, attempts, retry_delay)
        return future.result()


def _run_in_excel_impl(xlsx_path: Path, action, refresh_pivot: bool, attempts: int, retry_delay: float):
    """The actual implementation behind `_run_in_excel` — always called
    from a fresh, dedicated thread (see that function's docstring for
    why). Copies the file to a local temp dir — never the real file
    itself, since this workspace lives inside a OneDrive-synced folder,
    and Excel silently enables AutoSave for files stored there — opening
    the real deliverable via COM measurably rewrites it on close *even
    with SaveChanges=False* (observed: 91,083 -> 90,755 bytes, with
    parts dropped from the archive). The temp copy lives outside any
    synced folder, so the file the caller actually gets is never
    touched here.

    Opens the copy, forces a recalc, optionally refreshes the pivot
    table, then hands control to `action(excel, wb, sheet2)` to do the
    function-specific work and return its result. Closes the workbook
    and quits Excel afterward either way.

    Retries the *whole* attempt (a fresh Excel process each time) up to
    `attempts` times on the specific transient COM errors
    RPC_E_CALL_REJECTED / RPC_E_SERVERCALL_RETRYLATER ("Call was rejected
    by callee") — Excel intermittently rejects automation calls made
    right after Dispatch/Open, before it's fully settled. Any other
    error is raised immediately, no retry.

    Uses plain (late-bound) `win32com.client.Dispatch` rather than
    `gencache.EnsureDispatch` deliberately: the latter depends on a
    generated-wrapper cache under the user's temp folder that can go
    stale/corrupt (hit this directly during testing — it broke with
    "module ... has no attribute 'MinorVersion'" until the cache was
    cleared) and buys nothing here, since nothing in this module uses
    named Excel constants.
    """
    import pythoncom
    import pywintypes
    import win32com.client as win32
    import win32process

    for attempt in range(1, attempts + 1):
        # Callers may run this on a background thread; COM requires each
        # thread that touches it to initialize its own apartment first,
        # or Workbooks.Open can fail.
        pythoncom.CoInitialize()
        excel = None
        excel_pid = None
        tmp_dir = Path(tempfile.mkdtemp(prefix="workbook_com_"))
        tmp_copy = tmp_dir / xlsx_path.name
        try:
            shutil.copy2(xlsx_path, tmp_copy)

            excel = win32.Dispatch("Excel.Application")
            try:
                # Captured so cleanup can force-close *this specific*
                # spawned instance if Quit() doesn't fully take — a
                # leftover unresponsive EXCEL.EXE from one attempt was
                # observed to cause the very next attempt to fail the
                # same way, cascading retries into a solid wall of
                # rejected calls. Never touches any other Excel process
                # (e.g. one the user has open themselves).
                excel_pid = win32process.GetWindowThreadProcessId(excel.Hwnd)[1]
            except Exception:
                pass
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(str(tmp_copy))
            if wb is None:
                # Seen during testing: under the same contention that
                # causes RPC_E_CALL_REJECTED, Workbooks.Open can return
                # None instead of raising. Route it through the same
                # transient-retry path rather than failing with a
                # confusing "'NoneType' has no attribute ..." later.
                raise pywintypes.com_error(
                    -2147418111, "Workbooks.Open returned no workbook (transient)", None, None
                )
            try:
                wb.AutoSaveOn = False
            except Exception:
                pass
            excel.CalculateFullRebuild()

            sheet2 = wb.Sheets("Sheet2")
            if refresh_pivot:
                sheet2.PivotTables(1).RefreshTable()
                excel.CalculateFullRebuild()

            result = action(excel, wb, sheet2)
            wb.Close(SaveChanges=False)
            return result
        except pywintypes.com_error as e:
            transient = bool(e.args) and e.args[0] in _TRANSIENT_COM_HRESULTS
            if transient and attempt < attempts:
                time.sleep(retry_delay)
                continue
            raise
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
                # A brief cooldown after quitting: back-to-back COM
                # sessions (e.g. a retry, or a second tab/button click
                # moments after the last one closed) reliably triggered
                # "Call was rejected by callee" during testing without
                # this — Excel/COM seems to want a beat after one
                # process fully releases before the next activates
                # cleanly.
                time.sleep(0.5)
                if excel_pid and _process_alive(excel_pid):
                    _force_kill(excel_pid)
                    time.sleep(0.3)
            pythoncom.CoUninitialize()
            shutil.rmtree(tmp_dir, ignore_errors=True)


def _validate_in_excel_direct(path: Path, log) -> None:
    """The actual validate_in_excel implementation — only ever called
    from a fresh worker subprocess (see validate_in_excel below for why).
    Opens a *throwaway copy* of the generated file in real Excel, forces
    a full recalc and a pivot refresh, and reports any formula errors
    via `log(message)`. Degrades gracefully (skips with a note) if
    Excel/pywin32 isn't available. See `_run_in_excel` for the OneDrive/
    AutoSave-safety and transient-COM-retry details.
    """
    try:
        import pythoncom  # noqa: F401
        import win32com.client as win32  # noqa: F401
    except ImportError:
        log("Skipped Excel validation (pywin32 not installed).")
        return

    def action(excel, wb, sheet2):
        data_ws = wb.Sheets("Data")
        last_row = data_ws.UsedRange.Rows.Count

        errors = []
        for col in ("B", "C", "D"):
            values = data_ws.Range(f"{col}2:{col}{last_row}").Value
            for i, row in enumerate(values, start=2):
                v = row[0]
                if isinstance(v, str) and v.startswith("#"):
                    errors.append(f"{col}{i}: {v}")

        if errors:
            log(f"[WARN] {len(errors)} formula error(s) found after recalculation:")
            for e in errors[:15]:
                log(f"   {e}")
            if len(errors) > 15:
                log(f"   ... and {len(errors) - 15} more")
        else:
            log("[OK] No formula errors (#N/A / #NAME? / #VALUE!) after recalculation.")

        # Independent try/except (not `refresh_pivot=True`) so a pivot
        # refresh failure doesn't hide the formula-error results above.
        try:
            sheet2.PivotTables(1).RefreshTable()
            excel.CalculateFullRebuild()
            log("[OK] Pivot table refreshed successfully.")
        except Exception as e:
            log(f"[WARN] Could not refresh pivot table: {e}")

    try:
        _run_in_excel(path, action, refresh_pivot=False)
    except Exception as e:
        log(f"[WARN] Excel validation could not complete: {e}")


def _excel_cell_to_display(value):
    """Normalizes a value read back from Excel (COM can hand back
    pywintypes datetimes, floats, None, etc.) into something simple to
    render in an HTML table."""
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        # Whole-day pivot column headers show as dates; anything with a
        # time component keeps it.
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        # Pivot counts come back from COM as floats even though they're
        # always whole numbers — show "494" not "494.0".
        return int(value)
    return value


def _read_excel_pivot_direct(xlsx_path: Path) -> dict:
    """The actual read_excel_pivot implementation — only ever called from
    a fresh worker subprocess (see read_excel_pivot below for why). Opens
    a throwaway copy in real Excel, refreshes the pivot table, and reads
    back Sheet2's actual cell grid exactly as Excel renders it — the
    literal pivot output, not `compute_pivot`'s Python reproduction. See
    `_run_in_excel` for the OneDrive/AutoSave-safety and transient-COM-
    retry details. Returns {"ok": True, "rows": [[...], ...]} or
    {"ok": False, "error": ...}.
    """
    try:
        import pythoncom  # noqa: F401
        import win32com.client as win32  # noqa: F401
    except ImportError:
        return {"ok": False, "error": "Excel (pywin32) is not available on this machine."}

    def action(excel, wb, sheet2):
        raw = sheet2.UsedRange.Value
        if raw is None:
            return []
        if isinstance(raw, tuple) and raw and isinstance(raw[0], tuple):
            return [[_excel_cell_to_display(c) for c in r] for r in raw]
        if isinstance(raw, tuple):
            return [[_excel_cell_to_display(c) for c in raw]]
        return [[_excel_cell_to_display(raw)]]

    try:
        rows = _run_in_excel(xlsx_path, action, refresh_pivot=True)
        return {"ok": True, "rows": rows}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _copy_excel_pivot_to_clipboard_direct(xlsx_path: Path) -> dict:
    """The actual copy_excel_pivot_to_clipboard implementation — only
    ever called from a fresh worker subprocess (see
    copy_excel_pivot_to_clipboard below for why). Opens a throwaway copy
    in real Excel, refreshes the pivot table, and has Excel itself copy
    Sheet2's used range — giving an exact match of Excel's own rendering
    (cell colors, bold subtotal rows, borders, blank-vs-zero, the works)
    on paste, since it genuinely is Excel doing the copy rather than a
    reconstruction.

    Excel puts a rich set of formats on the clipboard while it's still
    running (confirmed: HTML Format, RTF, native Biff, CSV, ...) but only
    flushes a couple of them (RTF, plain text) once the process actually
    quits — "HTML Format" is what most paste targets (Outlook, Word,
    Teams, web apps) prefer for rich content, and without it they were
    silently falling back to plain text. So this reads HTML Format + RTF
    + plain text back *while Excel is still open* (right after the copy),
    closes Excel, then re-writes those exact bytes onto the clipboard
    itself via win32clipboard — confirmed this makes "HTML Format"
    reliably survive after Excel's process is gone.

    Paste with Ctrl+V into Excel/Outlook/Word/Teams. See `_run_in_excel`
    for the OneDrive/AutoSave-safety and transient-COM-retry details.
    Returns {"ok": True} or {"ok": False, "error": ...}.
    """
    try:
        import win32clipboard
    except ImportError:
        return {"ok": False, "error": "Excel (pywin32) is not available on this machine."}

    def action(excel, wb, sheet2):
        sheet2.UsedRange.Copy()
        time.sleep(0.3)  # let Excel finish rendering clipboard formats

        html_fmt = win32clipboard.RegisterClipboardFormat("HTML Format")
        rtf_fmt = win32clipboard.RegisterClipboardFormat("Rich Text Format")
        win32clipboard.OpenClipboard()
        try:
            html_bytes = win32clipboard.GetClipboardData(html_fmt)
            rtf_bytes = win32clipboard.GetClipboardData(rtf_fmt)
            text_data = win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
        finally:
            win32clipboard.CloseClipboard()
        return (html_fmt, rtf_fmt, html_bytes, rtf_bytes, text_data)

    try:
        html_fmt, rtf_fmt, html_bytes, rtf_bytes, text_data = _run_in_excel(xlsx_path, action, refresh_pivot=True)
    except Exception as e:
        return {"ok": False, "error": str(e)}

    # Re-own the clipboard ourselves with the captured bytes, now that
    # Excel (closed inside _run_in_excel) is fully gone — this is the
    # part that actually persists.
    win32clipboard.OpenClipboard()
    try:
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardData(html_fmt, html_bytes)
        win32clipboard.SetClipboardData(rtf_fmt, rtf_bytes)
        win32clipboard.SetClipboardData(win32clipboard.CF_UNICODETEXT, text_data)
    finally:
        win32clipboard.CloseClipboard()

    return {"ok": True}


# --- Subprocess worker: the real fix for RPC_E_CALL_REJECTED -----------
#
# Root-caused during development: within one long-lived process (this is
# exactly how the Flask dev server runs — single process, handling every
# request), the *first* Excel COM automation call always succeeded, and
# *every* call after that — even from a brand-new dedicated thread — then
# reliably failed with "Call was rejected by callee", regardless of
# delays or retry counts. Every genuinely separate process invocation
# succeeded, 100% reproducibly across many trials. So each of the three
# public functions below spawns this same script as a one-shot
# subprocess to do the actual Excel work, guaranteeing a pristine
# process every time — cheap relative to Excel automation's own runtime,
# and the only approach that's been fully reliable.
#
# The subprocess writes its JSON result to a temp file (passed as an
# argument) rather than stdout, sidestepping any buffering/encoding
# edge cases with large pivot payloads.

def _run_excel_worker(action: str, xlsx_path: Path) -> dict:
    """Runs one Excel-COM action ("validate" / "read_pivot" /
    "copy_clipboard") in a fresh subprocess and returns its JSON result.
    `copy_clipboard`'s effect (writing to the Windows clipboard) is
    system-wide and outlives the subprocess, same as it would in-process.
    """
    result_path = Path(tempfile.gettempdir()) / f"excel_worker_result_{os.getpid()}_{id(object())}.json"
    try:
        proc = subprocess.run(
            [sys.executable, str(Path(__file__).resolve()), "--excel-worker", action, str(xlsx_path), str(result_path)],
            capture_output=True, text=True, timeout=150,
        )
        if result_path.exists():
            with open(result_path, encoding="utf-8") as f:
                return json.load(f)
        stderr_tail = (proc.stderr or "").strip()
        return {
            "ok": False,
            "error": stderr_tail[-2000:] if stderr_tail else f"Excel worker exited with code {proc.returncode} and no result.",
        }
    except subprocess.TimeoutExpired:
        return {"ok": False, "error": "Excel automation timed out."}
    finally:
        result_path.unlink(missing_ok=True)


def validate_in_excel(path: Path, log) -> None:
    """Public entry point: runs `_validate_in_excel_direct` in a fresh
    worker subprocess (see the block comment above) and replays its log
    lines through `log(message)`, preserving the original calling
    convention."""
    result = _run_excel_worker("validate", path)
    for line in result.get("log", []):
        log(line)
    if "error" in result:
        log(f"[WARN] Excel validation could not complete: {result['error']}")


def read_excel_pivot(xlsx_path: Path) -> dict:
    """Public entry point: runs `_read_excel_pivot_direct` in a fresh
    worker subprocess (see the block comment above). Returns
    {"ok": True, "rows": [[...], ...]} or {"ok": False, "error": ...}."""
    return _run_excel_worker("read_pivot", xlsx_path)


def copy_excel_pivot_to_clipboard(xlsx_path: Path) -> dict:
    """Public entry point: runs `_copy_excel_pivot_to_clipboard_direct`
    in a fresh worker subprocess (see the block comment above). Returns
    {"ok": True} or {"ok": False, "error": ...}."""
    return _run_excel_worker("copy_clipboard", xlsx_path)


def _excel_worker_entrypoint(argv: list[str]) -> None:
    """Invoked as `py generate_workbook.py --excel-worker <action> <xlsx_path> <result_json_path>`
    — does exactly one Excel-COM action and writes its JSON result to
    the given path, then exits. See the block comment above `_run_excel_worker`."""
    action, xlsx_path_str, result_path_str = argv[0], argv[1], argv[2]
    xlsx_path = Path(xlsx_path_str)
    result_path = Path(result_path_str)

    if action == "validate":
        logs: list[str] = []
        _validate_in_excel_direct(xlsx_path, logs.append)
        result = {"log": logs}
    elif action == "read_pivot":
        result = _read_excel_pivot_direct(xlsx_path)
    elif action == "copy_clipboard":
        result = _copy_excel_pivot_to_clipboard_direct(xlsx_path)
    else:
        result = {"ok": False, "error": f"Unknown --excel-worker action: {action!r}"}

    with open(result_path, "w", encoding="utf-8") as f:
        json.dump(result, f)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--technical", required=True, type=Path)
    parser.add_argument("--functional", required=True, type=Path)
    parser.add_argument(
        "--template",
        required=True,
        type=Path,
        help="Most recent day's workbook — its Match sheet (and pivot shell) carry forward as-is.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=OUTPUT_DIR / default_output_name(),
    )
    args = parser.parse_args()

    generate_workbook(args.technical, args.functional, args.template, args.output)


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--excel-worker":
        _excel_worker_entrypoint(sys.argv[2:])
    else:
        main()
